#!/usr/bin/env python3
"""판정 결과를 리포트로 만든다.

사용법:
    python3 scripts/report.py <작업폴더> [--out <출력폴더>] [--title "..."]

입력:  <작업폴더>/verdicts.jsonl,  <작업폴더>/docs.json
출력:  <출력폴더>/충돌점검_리포트.md,  <출력폴더>/충돌목록.xlsx
"""
import argparse
import json
from collections import Counter
from datetime import date
from pathlib import Path

TYPE_LABEL = {
    "C1": "직접모순",
    "C2": "수치·기준 불일치",
    "C3": "폐기 누락",
    "C4": "위계 위반",
    "C5": "적용범위 충돌",
    "C6": "정의 불일치",
    "C7": "참조 끊김",
}
SEV_ORDER = {"높음": 0, "중간": 1, "낮음": 2}
SEV_MARK = {"높음": "●", "중간": "◐", "낮음": "○"}


def load(work: Path):
    rows = []
    p = work / "verdicts.jsonl"
    with p.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return [r for r in rows if r.get("verdict") != "N"]


def cite(side):
    loc = side.get("locator") or ""
    return f"{side.get('filename') or side.get('doc_id')} {loc}".strip()


def build_md(findings, docs, title):
    sev = Counter(f.get("severity", "낮음") for f in findings)
    typ = Counter(f.get("verdict") for f in findings)
    L = []
    L.append(f"# {title}")
    L.append("")
    L.append(f"작성일: {date.today().isoformat()}  |  대상 문서 {len(docs)}건  |  확인된 충돌 {len(findings)}건")
    L.append("")
    L.append("## 요약")
    L.append("")
    L.append("| 심각도 | 건수 |")
    L.append("|---|---|")
    for s in ("높음", "중간", "낮음"):
        if sev.get(s):
            L.append(f"| {SEV_MARK[s]} {s} | {sev[s]} |")
    L.append("")
    L.append("| 유형 | 건수 |")
    L.append("|---|---|")
    for t, c in sorted(typ.items(), key=lambda x: -x[1]):
        L.append(f"| {t} {TYPE_LABEL.get(t, t)} | {c} |")
    L.append("")

    if sev.get("높음"):
        L.append("> 심각도 높음 항목은 상위 규범 위반이거나 대외 효력이 있는 문서가 걸린 사안입니다. "
                 "먼저 처리하십시오.")
        L.append("")

    L.append("## 확인된 충돌")
    L.append("")
    findings.sort(key=lambda f: (SEV_ORDER.get(f.get("severity", "낮음"), 3), f.get("verdict", "")))
    for i, f in enumerate(findings, 1):
        v = f.get("verdict")
        L.append(f"### {i}. [{SEV_MARK.get(f.get('severity','낮음'),'○')} {f.get('severity','')}] "
                 f"{TYPE_LABEL.get(v, v)} — {f.get('subject') or ''}")
        L.append("")
        L.append(f"- **A** {cite(f['a'])}")
        L.append(f"  > {f['a'].get('text','').strip()}")
        L.append(f"- **B** {cite(f['b'])}")
        L.append(f"  > {f['b'].get('text','').strip()}")
        L.append("")
        L.append(f"**충돌 내용**  {f.get('rationale','')}")
        L.append("")
        if f.get("prevailing"):
            L.append(f"**우선 적용**  {f['prevailing']} — {f.get('prevailing_reason','')}")
            L.append("")
        if f.get("action"):
            L.append(f"**조치 제안**  {f['action']}")
            L.append("")
        if f.get("draft_text"):
            L.append("**수정 문안 초안**")
            L.append("")
            L.append("```")
            L.append(f["draft_text"].strip())
            L.append("```")
            L.append("")
        if f.get("confidence") == "낮음" or f.get("verify"):
            L.append(f"*확인 필요*: {f.get('verify','판정 근거가 약합니다. 원문 확인이 필요합니다.')}")
            L.append("")
        L.append("---")
        L.append("")

    L.append("## 검토 방법")
    L.append("")
    L.append("각 항목은 기계 판정 결과이며 확정된 사실이 아닙니다. 원문을 열어 확인한 뒤 "
             "채택 여부를 결정하십시오. 이 리포트는 원본 문서를 수정하지 않습니다.")
    L.append("")
    L.append("판정에서 제외된 정상 사례: 적용범위가 분리된 경우, 상위 규범이 명시적으로 "
             "위임한 특칙, 경과규정에 따라 시한부로 병존하는 조항.")
    return "\n".join(L)


def build_xlsx(findings, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "충돌목록"
    cols = ["번호", "심각도", "유형", "주제", "A 출처", "A 내용", "B 출처", "B 내용",
            "충돌 내용", "우선 적용", "조치 제안", "확신도", "담당", "처리상태"]
    ws.append(cols)
    head_fill = PatternFill("solid", fgColor="1F3864")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    fills = {"높음": PatternFill("solid", fgColor="FCE4E4"),
             "중간": PatternFill("solid", fgColor="FFF4E0"),
             "낮음": PatternFill("solid", fgColor="F2F2F2")}
    for i, f in enumerate(findings, 1):
        ws.append([
            i, f.get("severity", ""), TYPE_LABEL.get(f.get("verdict"), f.get("verdict", "")),
            f.get("subject", ""), cite(f["a"]), f["a"].get("text", ""),
            cite(f["b"]), f["b"].get("text", ""), f.get("rationale", ""),
            f.get("prevailing", ""), f.get("action", ""), f.get("confidence", ""), "", "미처리",
        ])
        fill = fills.get(f.get("severity", "낮음"))
        if fill:
            ws.cell(row=ws.max_row, column=2).fill = fill
    widths = [6, 8, 16, 22, 24, 46, 24, 46, 46, 14, 40, 8, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("work")
    ap.add_argument("--out", default=None)
    ap.add_argument("--title", default="사내 문서 충돌 점검 리포트")
    args = ap.parse_args()

    work = Path(args.work)
    out = Path(args.out) if args.out else work
    out.mkdir(parents=True, exist_ok=True)

    findings = load(work)
    docs = json.loads((work / "docs.json").read_text(encoding="utf-8"))

    md_path = out / "충돌점검_리포트.md"
    md_path.write_text(build_md(findings, docs, args.title), encoding="utf-8")
    xlsx_path = out / "충돌목록.xlsx"
    build_xlsx(findings, xlsx_path)
    print(f"충돌 {len(findings)}건")
    print(f"→ {md_path}")
    print(f"→ {xlsx_path}")


if __name__ == "__main__":
    main()